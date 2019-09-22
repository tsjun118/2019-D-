"""华为杯全国大学生数学建模   WEKEEP团队    2019年9月23日
   该文件代码用于：第一问的数据预处理和第二问的运动学片段划分"""


"""汽车过隧道，信号被覆盖时，GPS信号异常监测"""
import numpy as np
from openpyxl import Workbook, load_workbook
import xlwt
import pandas as pd
import  csv



def save_to_csv(data, filename, headline=None):
    with open(filename, 'w', newline='') as f:
        if headline == None:
            writer = csv.writer(f)
            writer.writerows(data)
        else:
          writer = csv.writer(f)
          writer.writerow(headline)
          writer.writerows(data)



# 导入工作目录
filename =r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2\文件23\置gps异常值为0_文件3.xlsx'
file_write_excel = r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2'
wb = load_workbook(filename)
combine_sheet = wb['原始数据3']

# 获取数据
GPS, zhuansu,zhuansu_n0s = [], [], []
h = 164915  #选取的行数
for row in range(2, h):
    current_GPS = combine_sheet.cell(row=row, column=2).value
    current_zhuansu = combine_sheet.cell(row=row, column=8).value
    current_youhao = combine_sheet.cell(row=row, column=10).value
    current_youmen = combine_sheet.cell(row=row, column=11).value
    zhuansu.append(current_zhuansu)
    # GPS.append(current_GPS)

    if  current_GPS == 0  and current_youmen == 0 :
        zhuansu_n0 = current_zhuansu
        zhuansu_n0s.append(zhuansu_n0)     #计算n0转速的列表
    if current_GPS != 0 or current_youhao != 0:    #排除出车速为0和油耗为0的值
        GPS.append(current_GPS)                    #这个列表数据就是最终数据

n0 = np.mean(zhuansu_n0s)    #怠速时的转速
print("怠速时转速：")
print(n0)



# file_write_csv = r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2\gps_xg2.csv'
# save_to_csv(GPS, file_write_csv)
# csvfile = open('gsp_xg2.csv', 'w', newline='')
# writer = csv.writer(csvfile)
# m = len(GPS)
# for i in range(m):
#     writer.writerows(GPS)
#     writer.writerlists(GPS)
# csvfile.close()



"""检测gps异常值并输出"""
i = 0
yichang_rows = []
for gps in GPS:
    if gps == 0:
        k = zhuansu[i]
        if k > n0:
            yichang_row = i + 2
            yichang_rows.append(yichang_row)    #异常行数的列表
    i = i+1



# for del_row in yichang_rows:
#     combine_sheet.cell(row=del_row, column=2, value='')
# wb.save(r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2\文件23\置gps异常值为0_文件3.xlsx')
#
# print('GPS异常行数：')
# print(yichang_rows)

# l = len(GPS)-1
# path = "./删掉熄火的数据2.csv"
# with open(path, "w", newline="", encoding="utf-8") as f:
#      try:
#         csvWriter = csv.writer(f,dialect="excel")
#         csvWriter.writerows(GPS)
#      finally:
#         f.close()
#
#
#
file_write_csv = r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2\文件23\删掉熄火的数据_文件2.csv'
test = pd.DataFrame(data = GPS)
test.to_csv(file_write_csv, encoding='gbk')

"""删除GPS空值"""
GPS_del_none = []
for gps in GPS:
    if gps != None:
        GPS_del_none.append(gps)

"""低速置为怠速"""
for i in range(len(GPS_del_none)):
    try:
        h = float(str(GPS_del_none[i]))
    except ValueError:
        pass
    else:
        if h<5:
            GPS_del_none[i] = 0




"""运动学片段划分"""
GPS_fd = []
l = len(GPS_del_none)
for i in range(1,l):
    h0 = GPS_del_none[i-1]
    h1 = GPS_del_none[i]
    GPS_fd.append(h1)
    if h0 > h1 and h1 == 0:
        GPS_fd.append("sky")

csvfile = r'C:\Users\Administrator.SC-201808311724\Desktop\jianmo2\运动学片段划分\文件3_运动学片段划分.txt'
fenduanshu = 1
with open(csvfile, "a") as f:
    for shuju in GPS_fd:
        if shuju == "sky":
            f.write("\r\n")
            fenduanshu = fenduanshu + 1
        else:
            shuju_str = str(shuju)
            f.write(shuju_str)
            f.write(",")
    fenduanshu_str = str(fenduanshu)
    f.write("\r\n")
    f.write("运动学片段分段数：")
    f.write(fenduanshu_str)


print('over')
# print(GPS_fd)
# print(GPS_fd)





